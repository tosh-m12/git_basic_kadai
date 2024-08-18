import pandas as pd
from openpyxl import load_workbook
import os
import win32com.client
import re
import datetime
import mysql.connector
from pathlib import Path
import pythoncom

pythoncom.CoInitialize()

# MySQL database connection details
db_config = {
    'user': 'root',
    'password': 'ngls@123',
    'host': 'localhost',
    'database': 'dnmasterlist'
}

def connect_to_db():
    return mysql.connector.connect(**db_config)

# Load customer data
customers_df = pd.read_csv('invoicing/data/extracted_values.csv')

# Load the invoice template
template_path = 'invoicing/data/DNformatForvia.xlsx'

bill_mo = 202406
DNdate = datetime.date.today()
billper = "June 1 ~ 30, 2024"
exrate = 7.1268
exratedate = "06/28/2024"

# Function to sanitize the company name
def sanitize_filename(name):
    return re.sub(r'[^A-Za-z0-9]+', '_', name)

db_conn = connect_to_db()
cursor = db_conn.cursor()

# Create invoices for each customer
for index, customer in customers_df.iterrows():
    try:
        # Skip if "Grand total" is zero
        if customer['Grand total'] == 0:
            continue

        # Load the template workbook and select the "attachment" sheet
        wb = load_workbook(template_path)
        ws1 = wb["DN"]
        ws2 = wb["Attachment"]

        attn = ""
        address = ""
        tel = ""
        refnum = ""

        # Assign values to the template fields (this example assumes specific cells for each field)
        ws1['E10'] = customer['Company name']
        ws1['V10'] = DNdate
        ws1['E12'] = attn
        ws1['E14'] = address
        ws1['E16'] = tel
        ws1['V12'] = refnum
        ws1['J29'] = exratedate
        ws1['O29'] = exrate
        ws2['C9'] = customer['Company name']
        ws2['C11'] = DNdate
        ws2['C13'] = refnum
        ws2['D15'] = billper
        ws2['E20'] = customer['Total stock']
        ws2['I20'] = customer['IV value']
        ws2['G20'] = customer['EOM inventory']
        ws2['E23'] = customer['Total IN']
        ws2['G24'] = customer['Total OUT']
        ws2['E25'] = customer['Storage']
        ws2['E26'] = customer['Labelling']
        ws2['E27'] = customer['IMMEX en']
        ws2['G28'] = customer['IMMEX vex']
        ws2['I29'] = customer['Ins']
        ws2['I30'] = customer['Sub total']
        ws2['I31'] = customer['16% VAT']
        ws2['I32'] = customer['Grand total']

        # Sanitize the company name for the filename
        sanitized_company_name = sanitize_filename(customer['Company name'])

        # Save the invoice with a unique name
        output_dir = f"data/DN/{bill_mo}/{sanitized_company_name}/"

        # Ensure the output directory exists
        os.makedirs(output_dir, exist_ok=True)

        invoice_filename = f"{output_dir}{sanitized_company_name}.xlsx"
        pdf_filename = f"{output_dir}{sanitized_company_name}.pdf"
        wb.save(invoice_filename)

        # Use win32com to ensure data is properly saved and to convert to PDF
        excel = win32com.client.Dispatch("Excel.Application")
        abspath = str(Path(invoice_filename).resolve())
        wb = excel.Workbooks.Open(abspath)
        excel.Visible = False
        excel.DisplayAlerts = False
        ws = wb.Worksheets(1)
        wb.Save()

        # Save as PDF
        pdf_abspath = str(Path(pdf_filename).resolve())
        wb.ExportAsFixedFormat(0, pdf_abspath)
        
        wb.Close()
        excel.Quit()

        wb3 = load_workbook(invoice_filename, data_only=True)
        ws3 = wb3['DN']
        sto = ws3['V20'].value
        whs = ws3['V21'].value
        svc = ws3['V22'].value
        trs = ws3['V23'].value
        dsb = ws3['V24'].value
        oth = ws3['V25'].value
        vat6 = ws3['V26'].value
        vat9 = ws3['V27'].value
        refnum = "NGLSABC0001"
        ccode = "NGLSABC"
        bp_from = 20240703
        bp_to = 20240803
        entry_dt = 20240804

        print(sto, whs, svc, trs, dsb, oth, vat6, vat9)

        # Prepare values for insertion, skipping None values
        values = [refnum, ccode, bp_from, bp_to, sto, whs, svc, trs, dsb, oth, vat6, vat9, entry_dt]
        columns = ['refnum', 'ccode', 'bp_from', 'bp_to', 'sto', 'whs', 'svc', 'trs', 'dsb', 'oth', 'vat6', 'vat9', 'entry_dt']

        insert_columns = []
        insert_values = []
        for col, val in zip(columns, values):
            if val is not None:
                insert_columns.append(col)
                insert_values.append(val)

        if insert_columns:
            insert_query = f"INSERT INTO dnmasterlist ({', '.join(insert_columns)}) VALUES ({', '.join(['%s'] * len(insert_values))})"
            cursor.execute(insert_query, insert_values)

    except Exception as e:
        print(f"Error processing customer {customer['Company name']}: {e}")

# Commit the transaction and close the database connection
db_conn.commit()
cursor.close()
db_conn.close()

print("DN have been created successfully.")
