import os
import re
import pandas as pd
import qrcode
from qrcode.image.pil import PilImage
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from sqlalchemy import create_engine
from urllib.parse import quote
from PIL import Image as PILImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D

group_code = 'FAURECIA'

def connect_to_database():
    try:
        password = quote('ngls@123')  # URL-encode the password
        connection_string = f'mysql+mysqlconnector://root:{password}@localhost/dnmasterlist'
        engine = create_engine(connection_string)
        return engine
    except Exception as e:
        print(f"Error connecting to MySQL: {e}")
        return None

def extract_ids_from_bd_mexico(engine):
    try:
        query = "SELECT dnmasterlist_id, rev FROM bd_mexico WHERE `new` = '*' AND grand_total <> 0"
        df = pd.read_sql(query, engine)
        print("IDs extracted from bd_mexico")
        return df
    except Exception as e:
        print(f"Error extracting IDs from bd_mexico: {e}")
        return None

def extract_data(engine, ids):
    try:
        id_list = ids['dnmasterlist_id'].tolist()
        query = f"""
            SELECT
                dn.id, bill_from, bill_to, sto, whs, svc, trs, dsb, oth, cn_vat6, cn_vat9,
                rate_date, rate, sub_total_cny, sub_total_usd, iv_amount_cny, iv_amount_usd,
                d_issued, dn.c_code, ac_term, serial_num, cc.company_name AS customer_company_name,
                rl.add1, rl.add2, rl.pic, rl.mr_ms, rl.tel, rl.company_name_full
            FROM dnmasterlist dn
            JOIN customer_code cc ON dn.c_code = cc.c_code
            LEFT JOIN recipientlist rl ON dn.c_code = rl.c_code AND rl.b_cat = 'A'
            WHERE dn.id IN ({','.join(map(str, id_list))})
        """
        df = pd.read_sql(query, engine)
        df = df.merge(ids, left_on='id', right_on='dnmasterlist_id')
        return df
    except Exception as e:
        print(f"Error extracting data: {e}")
        return None

def sanitize_name(name):
    sanitized = re.sub(r'[^a-zA-Z0-9_-]', '_', name)  # Replace invalid characters with underscore
    sanitized = re.sub(r'_+', '_', sanitized)         # Replace multiple underscores with a single one
    sanitized = sanitized.strip('_')                  # Remove leading and trailing underscores
    return sanitized

def generate_qr_code(data, qr_path):
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=10,
        border=4,
    )
    qr.add_data(data)
    qr.make(fit=True)
    img = qr.make_image(fill_color='#1A005D', back_color='white', image_factory=PilImage)
    img = img.convert("RGB")  # Convert image to RGB mode
    img = img.resize((252, 252), PILImage.LANCZOS)  # Resize to 25.2 mm x 25.2 mm (1 mm = 10 pixels)
    img.save(qr_path)

def update_existing_bd(company_folder, sanitized_company_name, d_issued_formatted, combined_code, rev_suffix):
    try:
        base_filename = f"BD_{sanitized_company_name}{rev_suffix}"
        existing_file_base = os.path.join(company_folder, f"{base_filename}.xlsx")

        if os.path.exists(existing_file_base):
            book = load_workbook(existing_file_base)
            sheet = book.active
            sheet['C11'] = d_issued_formatted
            sheet['C13'] = combined_code
            book.save(existing_file_base)
            print(f"Updated existing file: {existing_file_base}")
        else:
            print(f"File not found: {existing_file_base}")
    except Exception as e:
        print(f"Error updating existing Excel file: {e}")

def write_to_excel(dataframe, template_path, base_output_path):
    try:
        if not dataframe.empty:
            for idx, row in dataframe.iterrows():
                unsanitized_company_name = row['customer_company_name']
                sanitized_company_name = sanitize_name(row['customer_company_name'])
                ac_term = sanitize_name(row['ac_term'])
                company_folder = os.path.join(base_output_path, ac_term, sanitized_company_name)
                os.makedirs(company_folder, exist_ok=True)

                base_filename = f"DN_{sanitized_company_name}"  # With 'DN_' prefix
                rev_suffix = f"_R{row['rev'][-2:]}" if row['rev'] else ""
                output_path = os.path.join(company_folder, f"{base_filename}{rev_suffix}.xlsx")

                book = load_workbook(template_path)
                sheet_name = 'DN'
                if sheet_name not in book.sheetnames:
                    print(f"Sheet {sheet_name} not found in the template.")
                    continue

                sheet = book[sheet_name]

                d_issued_formatted = row['d_issued'].strftime('%Y%m%d')
                combined_code = (row['c_code'][-6:] + '-' + row['ac_term'] + '-' + row['serial_num'])
                total_usd_before_tax = row['sto'] + row['whs'] + row['svc'] + row['trs'] + row['dsb'] + row['oth']

                sheet['V10'] = d_issued_formatted
                sheet['V12'] = combined_code
                sheet['V14'] = group_code
                sheet['T18'] = row['bill_from']
                sheet['Y18'] = row['bill_to']
                sheet['V22'] = row['sto']
                sheet['V23'] = row['whs']
                sheet['V24'] = row['svc']
                sheet['V25'] = row['trs']
                sheet['V26'] = row['dsb']
                sheet['V27'] = row['oth']
                sheet['V28'] = total_usd_before_tax
                sheet['V29'] = row['cn_vat6']
                sheet['V30'] = row['cn_vat9']
                sheet['J33'] = row['rate_date']
                sheet['O33'] = row['rate']
                sheet['V33'] = row['sub_total_cny']
                sheet['V31'] = row['sub_total_usd']
                sheet['V36'] = row['iv_amount_cny']
                sheet['V37'] = row['iv_amount_usd']
                sheet['E10'] = row['company_name_full']
                sheet['E12'] = f"{row['pic']} {row['mr_ms']}".strip() if row['mr_ms'] else row['pic']
                sheet['E14'] = row['add1'] if row['add1'] else ''
                sheet['E16'] = row['add2'] if row['add2'] else ''
                sheet['E18'] = row['tel'] if row['tel'] else ''

                qr_data = f"{combined_code},{d_issued_formatted}"
                qr_path = os.path.join(company_folder, f"QR_{sanitized_company_name}.png")
                generate_qr_code(qr_data, qr_path)
                img = Image(qr_path)
                img.width = img.height = 98

                def pixels_to_EMUs(pixels):
                    return int(pixels * 914400 / 72)  # 1 pixel = 914400 EMUs / 72 DPI

                col_offset = pixels_to_EMUs(0)
                row_offset = -pixels_to_EMUs(6)

                img.anchor = OneCellAnchor(
                    _from=AnchorMarker(col=24, colOff=col_offset, row=1, rowOff=row_offset),
                    ext=XDRPositiveSize2D(img.width * 9525, img.height * 9525)
                )

                sheet.add_image(img)
                book.save(output_path)
                print(f"Invoice created and saved as {output_path}")

                update_existing_bd(company_folder, sanitized_company_name, d_issued_formatted, combined_code, rev_suffix)
    except Exception as e:
        print(f"Error writing to Excel: {e}")

def main():
    engine = connect_to_database()
    if engine is not None:
        try:
            ids = extract_ids_from_bd_mexico(engine)
            if not ids.empty:
                df = extract_data(engine, ids)
                if df is not None and not df.empty:
                    write_to_excel(df, './Forvia_mailsys/invoicing/data/DNformatForvia.xlsx', './invoicing/data/')
                else:
                    print("No data extracted or data is empty.")
        finally:
            engine.dispose()

if __name__ == "__main__":
    main()
