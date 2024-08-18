import pandas as pd
import os
import re
from openpyxl import load_workbook
from sqlalchemy import create_engine, text
import urllib.parse

################ ADJUST PARAMETERS HERE ################

# MySQL connection details
mysql_host = os.getenv('MYSQL_HOST', 'localhost')
mysql_user = os.getenv('MYSQL_USER', 'root')
mysql_password = urllib.parse.quote_plus(os.getenv('MYSQL_PASSWORD', 'ngls@123'))
mysql_database = os.getenv('MYSQL_DATABASE', 'dnmasterlist')

# Source table
source_table = 'f_extract'
lookup_table = 'bd_mexico'
template_invoice_a = './Forvia_mailsys/invoicing/data/BDformatForvia.xlsx'
template_invoice_c = './Forvia_mailsys/invoicing/data/BDformatForvia_seat.xlsx'

########################################################

def get_sqlalchemy_engine(host, user, password, database):
    return create_engine(f'mysql+mysqlconnector://{user}:{password}@{host}/{database}')

def extract_ids_from_bd_mexico(engine):
    try:
        query = f"SELECT f_extract_id, rev FROM {lookup_table} WHERE `new` = '*' AND grand_total <> 0"
        df = pd.read_sql(query, engine)
        print("IDs extracted from bd_mexico")
        return df[['f_extract_id', 'rev']].to_dict(orient='records')
    except Exception as err:
        print(f"Error extracting IDs from bd_mexico: {err}")
        return None

def extract_data_from_f_extract(engine, ids):
    try:
        id_list = [str(item['f_extract_id']) for item in ids]
        query = f"SELECT * FROM {source_table} WHERE id IN ({','.join(id_list)})"
        df = pd.read_sql(query, engine)
        print("Data extracted successfully from f_extract")
        return df
    except Exception as err:
        print(f"Error extracting data from f_extract: {err}")
        return None

#def sanitize_filename(filename):
#    return "".join([c for c in filename if c.isalpha() or c.isdigit() or c==' ']).rstrip().replace(' ', '_')

def sanitize_filename(filename):
    sanitized = re.sub(r'[^a-zA-Z0-9_-]', '_', filename)  # Replace invalid characters with underscore
    sanitized = re.sub(r'_+', '_', sanitized)         # Replace multiple underscores with a single one
    sanitized = sanitized.strip('_')                  # Remove leading and trailing underscores
    return sanitized


def create_invoice(customer, rev_suffix):
    try:
        wb = load_workbook(template_invoice_a)
        ws = wb.active

        # Fill in the invoice fields
        ws['C9'] = customer['company_name']
        ws['G9'] = customer['c_code'][-6:]  # Take only the last 6 characters of c_code
        ws['E15'] = customer['bill_from']
        ws['G15'] = customer['bill_to']
        ws['E20'] = customer['total_stock']
        ws['G20'] = customer['eom_inventory']
        ws['I20'] = customer['iv_value']
        ws['E23'] = customer['total_in']
        ws['G24'] = customer['total_out']
        ws['E25'] = customer['storage']
        ws['E26'] = customer['labelling']
        ws['E27'] = customer['immex_en']
        ws['G28'] = customer['immex_vex']
        ws['I29'] = customer['ins']
        ws['I30'] = customer['sub_total']
        ws['I31'] = customer['16percent_vat']
        ws['I32'] = customer['grand_total']
        ws['I39'] = customer['grand_total']
        ws['I40'] = customer['col_chrg']
        ws['I41'] = customer['iv_amount']
        ws['I46'] = customer['cn_vat6']

        # Calculate sum of iv_amount and cn_vat6
        ws['I47'] = customer['iv_amount'] + customer['cn_vat6']

        # Generate sanitized filename
        sanitized_name = sanitize_filename(customer['company_name'])
        b_cat = customer['b_cat']
        bill_mo = customer['ac_term']
        output_dir = f"invoicing/data/{bill_mo}{b_cat}/{sanitized_name}/".replace(' ', '_')
        os.makedirs(output_dir, exist_ok=True)
        
        # Determine file name with revision suffix if necessary
        base_filename = f"{output_dir}BD_{sanitized_name}{rev_suffix}.xlsx"
        
        # Save the invoice, overwrite if it already exists
        wb.save(base_filename)
        print(f"Invoice created and saved as {base_filename}")
    except Exception as err:
        print(f"Error creating invoice for {customer['company_name']}: {err}")

# Get SQLAlchemy engine
engine = get_sqlalchemy_engine(mysql_host, mysql_user, mysql_password, mysql_database)
print("SQLAlchemy engine created")

# Extract f_extract_id and rev from bd_mexico where new is "*"
id_rev_list = extract_ids_from_bd_mexico(engine)

if id_rev_list:
    # Extract data from f_extract based on the extracted IDs
    ids = [{'f_extract_id': item['f_extract_id']} for item in id_rev_list]
    revs = {item['f_extract_id']: item['rev'] for item in id_rev_list}
    data = extract_data_from_f_extract(engine, ids)

    if data is not None:
        for index, customer in data.iterrows():
            rev_suffix = f"_{revs[customer['id']]}" if revs[customer['id']] else ""
            create_invoice(customer, rev_suffix)
