import openpyxl
import win32com.client
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import datetime

outlook = win32com.client.Dispatch("Outlook.Application")


today = datetime.date.today()
today_month = today.month
today_mo2digit = '{:0>2}'.format(today_month)
today_day = today.day
today_day2digit = '{:0>2}'.format(today_day)


reminderlist = openpyxl.load_workbook(r'./Forvia_mailsys/data/reminderlist/{}{}{}reminderlist.xlsx'.format(today.year, today_mo2digit, today_day2digit))
rml = reminderlist.active

contactlist = openpyxl.load_workbook(r'./Forvia_mailsys/data/contactlist.xlsx')
cntcl = contactlist.active


tmp_supalias = '{sup_alias}'
tmp_bmonth = '{billing_month}'
tmp_byear = '{billing_year}'
tmp_bcat = '{billing_category}'
tmp_obalancecny = '{outstanding_balance_CNY}'
tmp_ivissdate = '{invoice_issue_date}'
tmp_pterm = '{payment_term}'
tmp_ptfrom = '{payment_term_from}'
tmp_pduedate = '{payment_due_date}'
tmp_daysod = '{days_overdue}'


for rml_rows in rml.iter_rows(min_row=2):
    sup_code = rml_rows[0].value
    sup_alias = rml_rows[1].value
    bill_period = rml_rows[2].value
    if rml_rows[3].value == 'SF1':
        bill_cat = 'Sea Freight Cost'
    elif rml_rows[3].value == 'WH2':
        bill_cat = 'Warehouse Cost'

    bill_amntcny = rml_rows[4].value
    iv_issuedate = str(format(rml_rows[5].value, '%B %d, %Y'))
    pmterm = str(rml_rows[6].value)
    pmt_from = str(rml_rows[7].value)
    pmt_duedate = str(format(rml_rows[8].value, '%B %d, %Y'))
    days_overdue = str(rml_rows[9].value)
    ost_balancecny = '{:,}'.format(rml_rows[10].value) + ' CNY'

    mailbody_temp = open('mailbody.txt')
    data = mailbody_temp.read()
    mailbody_temp.close()

    mailbody = data.replace(tmp_supalias, sup_alias)\
        .replace(tmp_bmonth, format(bill_period, '%B'))\
            .replace(tmp_byear,format(bill_period, '%Y'))\
                .replace(tmp_bcat, bill_cat)\
                    .replace(tmp_obalancecny, ost_balancecny)\
                        .replace(tmp_ivissdate, iv_issuedate)\
                            .replace(tmp_pterm, pmterm)\
                                .replace(tmp_ptfrom, pmt_from)\
                                    .replace(tmp_pduedate, pmt_duedate)\
                                        .replace(tmp_daysod, days_overdue)
    
    for cntcl_row in cntcl.iter_rows(min_row=2):
        if sup_code == cntcl_row[1].value:
            addresslist = []
            for cntcl_col in cntcl_row[3:]:
                if cntcl_col.value != None:
                    addresslist.append(cntcl_col.value)
                address = "; ".join(addresslist)


            print(address)

            mail = outlook.CreateItem(0)
            mail.to = address
            mail.cc = "toshio.murayama@nipponexpress.com.cn"
            mail.subject = 'OVERDUE PAYMENT REMINDER: ' + sup_alias + format(bill_period, ' %B %Y ' + bill_cat + ' Billing')
            mail.bodyFormat = 2
            mail.ReplyRecipients.Add

            mail.HTMLbody = mailbody



            mail.display(True)




#mail.send()

 